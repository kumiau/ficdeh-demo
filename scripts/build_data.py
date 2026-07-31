#!/usr/bin/env python3
"""
Convierte los Excel de programación del 13° FICDEH (programacion xls/) en JSON
normalizado por ciudad (data/<slug>.json) + un índice (data/index.json).

Ver PLAN.md para el contexto: hay dos formatos de origen distintos:
  - "flat": una fila por función (Fecha, Hora, Lugar, ...). Es el formato de
    casi todas las ciudades salvo Bogotá.
  - "cartelera": bloques por sede pensados para lectura humana, con la hora
    en la primera fila de cada sesión y las siguientes filas heredándola. Es
    el único formato disponible para la programación principal de Bogotá.

Uso: python3 scripts/build_data.py
"""
import json
import re
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "programacion xls"
OUT = ROOT / "data"

FESTIVAL_YEAR = 2026
MONTHS = {
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12,
}

# Ciudades a generar. Cada fuente ya fue inspeccionada a mano (ver PLAN.md)
# para decidir si conviene el archivo plano (más rico: sinopsis, poster,
# redes) o el de cartelera (el único que existe para el grueso de Bogotá).
CITY_CONFIG = [
    {"slug": "bogota", "name": "Bogotá", "sources": [
        ("BOGOTÁ/Bogotá_Programación 13° FICDEH.xlsx", "cartelera"),
    ]},
    # PAZOSFERA es un programa nacional en centros de reclusión (Armenia,
    # Barranquilla, Cartagena, Cali, Manizales, Pereira, Tunja, Quibdó...),
    # no exclusivo de Bogotá, aunque el archivo esté en esa carpeta.
    {"slug": "pazosfera", "name": "Pazósfera (centros de reclusión)", "sources": [
        ("BOGOTÁ/Programación _ PAZOSFERA.xlsx", "flat"),
    ]},
    {"slug": "armenia", "name": "Armenia", "sources": [
        ("ARMENIA/ Programación _ ARMENIA.xlsx", "flat"),
    ]},
    {"slug": "barranquilla", "name": "Barranquilla", "sources": [
        ("BARRANQUILLA/Programación _ BARRANQUILLA.xlsx", "flat"),
    ]},
    {"slug": "cali", "name": "Cali", "sources": [
        ("CALI/Programación _ CALI.xlsx", "flat"),
    ]},
    {"slug": "cartagena", "name": "Cartagena", "sources": [
        ("CARTAGENA/Programación _ CARTAGENA.xlsx", "flat"),
    ]},
    {"slug": "ibague", "name": "Ibagué", "sources": [
        ("IBAGUÉ/Programación _ IBAGUE.xlsx", "flat"),
    ]},
    {"slug": "manizales", "name": "Manizales", "sources": [
        ("MANIZALES/Programación _ MANIZALES.xlsx", "flat"),
    ]},
    {"slug": "medellin", "name": "Medellín", "sources": [
        ("MEDELLÍN/Programación _ MEDELLÍN.xlsx", "flat"),
    ]},
    {"slug": "municipios", "name": "Municipios (Tenjo)", "sources": [
        ("MUNICIPIOS /Programación _ TENJO.xlsx", "flat"),
    ]},
    {"slug": "pereira", "name": "Pereira", "sources": [
        ("PEREIRA/Programación _ PEREIRA.xlsx", "flat"),
    ]},
    {"slug": "quibdo", "name": "Quibdó", "sources": [
        ("QUIBDÓ/Programación _ QUIBDÓ.xlsx", "flat"),
    ]},
    {"slug": "tunja", "name": "Tunja", "sources": [
        ("TUNJA/Programación _ TUNJA.xlsx", "flat"),
    ]},
]

# Archivos deliberadamente NO usados (quedan documentados para quien revise):
#  - "<Ciudad>_Programación 13° FICDEH.xlsx" y "Programación _ <SEDE>.xlsx"
#    de sedes específicas (LA TERTULIA, MANIZALES FICMA) duplican, con menos
#    datos, lo que ya está en el archivo plano de la ciudad.
#  - "Cinemateca de Bogotá_Programación...", "Programación _ BIBLIORED.xlsx"
#    y "Programación _ CASAS DE JUVENTUD.xlsx": se verificó (cruzando los
#    nombres de sede) que todas sus sedes ya están dentro del archivo
#    maestro "Bogotá_Programación...". Incluirlos duplicaba sesiones
#    (misma fecha/hora/sede repetida). Quedan como candidatos para una
#    futura pasada de enriquecimiento (traer su sinopsis/poster/trailer
#    sin duplicar sesiones), no para Fase 1.
#  - "PROGRAMACIÓN RETROSPECTIVA 13° FICDEH.xlsx" es una recopilación
#    cruzada de funciones que ya aparecen en los archivos por ciudad.
#  - Hojas "ORGANIZA", "FORMACIÓN", "AGRADECIMIENTOS", "CLÚSTER MIA",
#    "SALAS ASOCIADAS", "SUBA Y FONTANAR", "INVITADOS": no son programación
#    con fecha/hora fija (créditos, o catálogos de curaduría sin función
#    asignada), quedan fuera del alcance de una cartelera.

FLAT_FIELD_MAP = {
    "FECHA": "date", "HORA": "time", "LUGAR": "venue", "CATEGORIA": "category",
    "NOMBRE": "title", "NOMBRE DEL DIRECTOR/A": "director", "DURACION": "durationMin",
    "PAIS": "country", "ANO": "year", "SINOPSIS EN ESPANOL": "synopsisEs",
    "SINOPSIS EN INGLES": "synopsisEn", "PERFIL DEL DIRECTOR/A": "directorProfile",
    "REDES PELICULA": "filmSocial", "REDES DIRECTOR": "directorSocial",
    "LINK POSTER": "posterUrl", "TRAILER": "trailerUrl",
    "TEMATICA PRINCIPAL": "themePrimary", "TEMATICA SECUNDARIA": "themeSecondary",
    "KIT DE PRENSA": "pressKitUrl", "CALIFICACION PARA EL PUBLICO": "ageRating",
}
# LINK / CONTRASEÑA son credenciales privadas de proyecciones en línea
# (link + clave para jurados/prensa) — nunca deben llegar a un dato público.
FLAT_FIELD_SKIP = {"LINK", "CONTRASENA"}

CARTELERA_FIELD_MAP = {
    "HORA": "time", "NOMBRE PELICULA": "title", "DIRECTOR/A": "director",
    "DURACION": "durationMin", "DURACION (EN MINUTOS)": "durationMin",
    "PAIS": "country", "ANO": "year", "CATEGORIA": "category",
    "SALA": "sala", "Q&A": "qa", "ESTRENO NACIONAL": "premiere",
}

DAY_SHEET_RE = re.compile(r"(\d{1,2})\s+([A-Z]{3,4})$")
TIME_RE = re.compile(r"(\d{1,2}):(\d{2})\s*([ap])\.?\s*m", re.IGNORECASE)

warnings = []


def norm(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return text.upper()


def clean_line(value):
    """Colapsa saltos de línea/espacios en campos cortos (lugar, país, etc.)."""
    if value is None:
        return None
    # Algunos títulos se digitaron como número en el Excel (p. ej. la
    # película "1982"), y str(float) los vuelve "1982.0".
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).replace("\n", ", ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip(" ,")
    return text or None


def clean_block(value):
    """Para campos largos (sinopsis, perfil): solo recorta bordes."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_int(value):
    if value is None or value == "":
        return None
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return None


def time_sort_key(time_label):
    if not time_label:
        return 24 * 60
    m = TIME_RE.search(str(time_label))
    if not m:
        return 24 * 60
    hour, minute, ampm = int(m.group(1)), int(m.group(2)), m.group(3).lower()
    if ampm == "p" and hour != 12:
        hour += 12
    if ampm == "a" and hour == 12:
        hour = 0
    return hour * 60 + minute


def parse_sheet_date(sheet_name):
    m = DAY_SHEET_RE.search(norm(sheet_name))
    if not m:
        return None
    day, month_abbr = int(m.group(1)), m.group(2)[:3]
    month = MONTHS.get(month_abbr)
    if not month:
        return None
    return f"{FESTIVAL_YEAR:04d}-{month:02d}-{day:02d}"


# ---------------------------------------------------------------------------
# Parser "flat": una fila = una función de una película.
# ---------------------------------------------------------------------------

def parse_flat_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = "FICDEH" if "FICDEH" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    header = [norm(c.value) for c in ws[1]]
    field_by_col = {}
    for idx, label in enumerate(header):
        if not label or label in FLAT_FIELD_SKIP:
            continue
        field = FLAT_FIELD_MAP.get(label)
        if field:
            field_by_col[idx] = field

    if "date" not in field_by_col.values() or "time" not in field_by_col.values():
        warnings.append(f"[flat] {path.name}: no se encontraron columnas Fecha/Hora, se omite")
        return []

    sessions = {}
    order = []
    for row in ws.iter_rows(min_row=2):
        values = {}
        for idx, field in field_by_col.items():
            if idx < len(row):
                values[field] = row[idx].value

        date_val = values.get("date")
        if not hasattr(date_val, "isoformat"):
            continue  # filas de notas/instrucciones sin fecha real

        date_iso = date_val.date().isoformat() if hasattr(date_val, "date") else str(date_val)
        time_label = clean_line(values.get("time"))
        venue = clean_line(values.get("venue"))
        key = (date_iso, time_label, venue)

        if key not in sessions:
            session = {
                "date": date_iso,
                "time": time_label,
                "timeSortKey": time_sort_key(time_label),
                "venueName": venue,
                "venueAddress": None,
                "films": [],
                "notes": [],
            }
            sessions[key] = session
            order.append(key)

        film = {
            "title": clean_line(values.get("title")),
            "director": clean_line(values.get("director")),
            "durationMin": to_int(values.get("durationMin")),
            "country": clean_line(values.get("country")),
            "year": to_int(values.get("year")),
            "category": clean_line(values.get("category")),
            "synopsisEs": clean_block(values.get("synopsisEs")),
            "synopsisEn": clean_block(values.get("synopsisEn")),
            "directorProfile": clean_block(values.get("directorProfile")),
            "filmSocial": clean_line(values.get("filmSocial")),
            "directorSocial": clean_line(values.get("directorSocial")),
            "posterUrl": clean_line(values.get("posterUrl")),
            "trailerUrl": clean_line(values.get("trailerUrl")),
            "themePrimary": clean_line(values.get("themePrimary")),
            "themeSecondary": clean_line(values.get("themeSecondary")),
            "pressKitUrl": clean_line(values.get("pressKitUrl")),
            "ageRating": clean_line(values.get("ageRating")),
        }
        if film["title"]:
            sessions[key]["films"].append(film)

    return [sessions[k] for k in order if sessions[k]["films"]]


# ---------------------------------------------------------------------------
# Parser "cartelera": bloques por sede, con la hora en la primera fila de
# cada sesión. Ver PLAN.md para la explicación del layout.
# ---------------------------------------------------------------------------

def is_header_row(cells):
    """cells: lista de (col_idx, valor) no vacíos de la fila."""
    for i, (col, val) in enumerate(cells):
        if norm(val) == "HORA":
            for _, val2 in cells[i + 1:]:
                if norm(val2).startswith("NOMBRE"):
                    return col
            return None
    return None


def build_field_map(row_cells, hora_col):
    field_by_col = {}
    for col, val in row_cells:
        if col < hora_col:
            continue
        field = CARTELERA_FIELD_MAP.get(norm(val))
        if field:
            field_by_col[col] = field
    return field_by_col


def resolve_labels(labels):
    """labels: textos sueltos sobre un header, en orden de arriba hacia abajo
    (el más cercano al header queda último). 1 label -> nombre de sede.
    2 -> [sede, dirección]. 3 -> [ciudad, sede, dirección]."""
    venue = address = city_label = None
    if len(labels) == 1:
        venue = labels[0]
    elif len(labels) == 2:
        venue, address = labels
    elif len(labels) >= 3:
        city_label, venue, address = labels[-3:]
    return venue, address, city_label


def finalize_block(block, blocks):
    if block is None:
        return
    field_by_col = block["field_by_col"]
    hora_col = block["hora_col"]
    title_col = field_by_col_key(field_by_col, "title")
    duration_col = field_by_col_key(field_by_col, "durationMin")
    # Columna sin encabezado justo después del título: en algunas hojas el
    # título quedó digitado ahí por error de captura en el Excel original.
    spare_title_col = title_col + 1 if title_col is not None and (title_col + 1) not in field_by_col else None

    current_session = None
    for row_cells in block["raw_rows"]:
        by_col = dict(row_cells)
        title = clean_line(by_col.get(title_col))
        if not title and spare_title_col is not None:
            title = clean_line(by_col.get(spare_title_col))
        duration = to_int(by_col.get(duration_col))
        time_val = by_col.get(hora_col)

        is_film = bool(title) and duration is not None
        if not is_film:
            note = " - ".join(str(v).strip() for _, v in row_cells if v not in (None, ""))
            if current_session is not None and note:
                current_session["notes"].append(note)
            continue

        if time_val not in (None, "") or current_session is None:
            time_label = clean_line(time_val) or (current_session["time"] if current_session else None)
            current_session = {
                "date": block["date"],
                "time": time_label,
                "timeSortKey": time_sort_key(time_label),
                "venueName": block["venue"],
                "venueAddress": block["address"],
                "films": [],
                "notes": [],
            }
            blocks.append(current_session)

        film = {"title": title, "durationMin": duration}
        for col, field in field_by_col.items():
            if field in ("title", "durationMin"):
                continue
            val = by_col.get(col)
            film[field] = to_int(val) if field == "year" else clean_line(val)
        current_session["films"].append(film)


def field_by_col_key(field_by_col, field_name):
    for col, field in field_by_col.items():
        if field == field_name:
            return col
    return None


def parse_cartelera_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sessions = []
    for sheet_name in wb.sheetnames:
        date_iso = parse_sheet_date(sheet_name)
        if not date_iso:
            continue
        ws = wb[sheet_name]
        max_col = min(ws.max_column, 30)

        pending_labels = []
        blank_run = 0
        block = None

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col):
            cells = [(c.column, c.value) for c in row if c.value not in (None, "")]
            n = len(cells)

            if n == 0:
                blank_run += 1
                continue

            hora_col = is_header_row(cells)
            if hora_col is not None:
                finalize_block(block, sessions)
                venue, address, _city = resolve_labels(pending_labels)
                block = {
                    "date": date_iso,
                    "venue": venue,
                    "address": address,
                    "hora_col": hora_col,
                    "field_by_col": build_field_map(cells, hora_col),
                    "raw_rows": [],
                }
                pending_labels = []
                blank_run = 0
                continue

            if n == 1 and blank_run > 0:
                pending_labels.append(clean_line(cells[0][1]))
                if len(pending_labels) > 3:
                    pending_labels.pop(0)
                blank_run = 0
                continue

            if block is not None:
                block["raw_rows"].append(cells)
            pending_labels = []
            blank_run = 0

        finalize_block(block, sessions)
    return sessions


# ---------------------------------------------------------------------------
# Índice derivado por película: agrupa las funciones de todas las ciudades
# por título de película, para poder "buscar película -> ver dónde y cuándo
# se proyecta" sin tener que recorrer las sesiones de cada ciudad.
# ---------------------------------------------------------------------------

FILM_META_FIELDS = [
    "director", "durationMin", "country", "year", "category", "sala", "qa", "premiere",
    "synopsisEs", "synopsisEn", "directorProfile", "filmSocial", "directorSocial",
    "posterUrl", "trailerUrl", "themePrimary", "themeSecondary", "pressKitUrl", "ageRating",
]


def film_group_key(title):
    """Clave de agrupación insensible a acentos/mayúsculas/espacios repetidos.
    Ver data/README.md: los mismos títulos vienen escritos distinto según la
    ciudad/fuente (p. ej. "Floresmiro " con espacio final)."""
    return re.sub(r"\s+", " ", norm(title)).strip()


# Mismo título escrito distinto según la fuente (con/sin subtítulo en
# inglés o español, "/" en vez de paréntesis, un typo puntual). Se detectó
# revisando a mano data/films.json después de generarlo la primera vez —
# no es un algoritmo de coincidencia difusa (riesgo de fusionar películas
# distintas), es una lista curada. Si al regenerar aparecen títulos nuevos
# sin fusionar, hay que revisar data/films.json y sumarlos aquí.
TITLE_ALIASES = {
    "La cerillana": "La Cerrillana",
    "In four stops (En cuatro paradas)": "In Four Stops",
    "FLORES MIRO": "Floresmiro",
    "Feito Pipa /Gugu's world": "El mundo de Gugu (Feito Pipa)",
    "My grandmother is a skydiver (Mi abuela es una paracaidista)": "My grandmother is a skydiver",
    "The beauty of the donkey": "La Belleza del burro",
    "La belleza del burro (The beauty of the donkey)": "La Belleza del burro",
    "Three black men (Tres hombres negros)": "Three black men",
    "Padamlágan (Night light)": "Night Light",
    "Cuando la palabra se hace búsqueda. El eco de sus voces (UBPD)": "Cuando la palabra se hace búsqueda",
}
TITLE_ALIASES_BY_KEY = {
    re.sub(r"\s+", " ", norm(variant)).strip(): canonical
    for variant, canonical in TITLE_ALIASES.items()
}


def slugify(title):
    slug = re.sub(r"[^a-z0-9]+", "-", norm(title).lower()).strip("-")
    return slug or "sin-titulo"


def build_film_index(city_sessions):
    """city_sessions: lista de (citySlug, cityName, sessions) por ciudad ya
    generados. Cuando dos apariciones del mismo título traen metadata
    distinta (p. ej. el archivo cartelera de Bogotá no trae sinopsis pero el
    plano de otra ciudad sí), se completan los campos vacíos con el primer
    valor no vacío que aparezca — no se sobreescribe lo ya encontrado."""
    films = {}
    order = []
    for city_slug, city_name, sessions in city_sessions:
        for session in sessions:
            for film in session["films"]:
                title = film.get("title")
                if not title:
                    continue
                raw_key = film_group_key(title)
                title = TITLE_ALIASES_BY_KEY.get(raw_key, title)
                key = film_group_key(title)
                if key not in films:
                    entry = {"filmKey": slugify(title), "title": title, "screenings": []}
                    for field in FILM_META_FIELDS:
                        entry[field] = film.get(field)
                    films[key] = entry
                    order.append(key)
                else:
                    entry = films[key]
                    for field in FILM_META_FIELDS:
                        if entry.get(field) in (None, "") and film.get(field) not in (None, ""):
                            entry[field] = film.get(field)

                films[key]["screenings"].append({
                    "citySlug": city_slug,
                    "cityName": city_name,
                    "date": session["date"],
                    "time": session["time"],
                    "timeSortKey": session["timeSortKey"],
                    "venueName": session["venueName"],
                    "venueAddress": session["venueAddress"],
                })

    result = [films[k] for k in order]
    for film in result:
        film["screenings"].sort(key=lambda s: (s["date"] or "", s["timeSortKey"]))
    result.sort(key=lambda f: film_group_key(f["title"]))
    return result


# ---------------------------------------------------------------------------

def main():
    OUT.mkdir(exist_ok=True)
    index = {"festivalYear": FESTIVAL_YEAR, "cities": [], "warnings": warnings}
    city_sessions_for_index = []

    for city in CITY_CONFIG:
        all_sessions = []
        source_files = []
        for rel_path, kind in city["sources"]:
            path = SRC / rel_path
            if not path.exists():
                warnings.append(f"[{city['slug']}] archivo no encontrado: {rel_path}")
                continue
            source_files.append(rel_path)
            if kind == "flat":
                all_sessions.extend(parse_flat_file(path))
            else:
                all_sessions.extend(parse_cartelera_file(path))

        all_sessions.sort(key=lambda s: (s["date"] or "", s["timeSortKey"], s["venueName"] or ""))
        film_count = sum(len(s["films"]) for s in all_sessions)

        out_path = OUT / f"{city['slug']}.json"
        out_path.write_text(
            json.dumps({
                "citySlug": city["slug"],
                "cityName": city["name"],
                "sourceFiles": source_files,
                "sessions": all_sessions,
            }, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        index["cities"].append({
            "slug": city["slug"],
            "name": city["name"],
            "sessionCount": len(all_sessions),
            "filmCount": film_count,
        })
        city_sessions_for_index.append((city["slug"], city["name"], all_sessions))
        print(f"{city['name']}: {len(all_sessions)} sesiones, {film_count} películas -> {out_path.relative_to(ROOT)}")

    films = build_film_index(city_sessions_for_index)
    (OUT / "films.json").write_text(
        json.dumps({"films": films}, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    index["filmIndexCount"] = len(films)
    print(f"Índice por película: {len(films)} títulos únicos -> {(OUT / 'films.json').relative_to(ROOT)}")

    (OUT / "index.json").write_text(
        json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if warnings:
        print("\nAdvertencias:")
        for w in warnings:
            print(f"  - {w}")


if __name__ == "__main__":
    main()
